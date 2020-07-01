import numpy as np
import pandas as pd
import scipy.stats as ss
import xlrd
import openpyxl
import os
import sklearn
from datetime import timedelta, datetime
from dateutil.relativedelta import relativedelta
import time
import re
import pdb
from collections import OrderedDict, namedtuple
import pickle

os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio")
import defs
import cme_scrapper

#% import Data from Com pricing spreadsheet
#%Data Helper Functions
def int_col(n,base=26):
    "convert 0 index column number to excel column name"
    if n < 0:
        return ""
    elif n < base:
        return chr(65 + n)
    return int_col(n//base-1,base) + int_col(n%base,base)

def make_blb_book_row_aligned(file_name):
    """Takes file name, of (Bloomberg formatted security columns),
    Writes book to make all sheet's securities's columns date aligned.
    iterates over worksheets calling make_blb_sheet_aligned
    """
    bk_open = False
    xlrd_bk = xlrd.open_workbook(file_name)

    for sht_ix, name in enumerate(xlrd_bk.sheet_names()):
        b = xlrd_bk.sheet_by_index(sht_ix)
        sz = len([i for i in b.row_values(6)])
        #See blb_dates_are_skipped()
        date_cols = [[datetime(*xlrd.xldate_as_tuple(j,0))
                        for j in b.col_values(i)[7:]
                    if j]
                        for i in range(0,sz,4)]
        dates = sorted(set([i
                            for c in date_cols
                                for i in c]),
                        reverse = True)
        is_date_aligned = [col[-1] == dates[len(col)-1]
                               for col in date_cols]

        if not all(is_date_aligned):
            print(f"\nWARNING: Row's not date aligned, some days missing, for {name}")
            bad_date_cols = [4*ix for ix,i in enumerate(is_date_aligned)
                                if not i]
            print(bad_date_cols)
            if not bk_open:
                wrote_bk = openpyxl.load_workbook(file_name, data_only = True)
                bk_open = True
            #keeps book open in case have to write to another sheet; opening takes forever
            wrote_bk = make_blb_sheet_aligned(file_name,
                                                 sht_ix,
                                                 bad_date_cols,
                                                 dates,
                                                 wrote_bk)
    if bk_open:
        wrote_bk.save(file_name + "(2)")
        wrote_bk.close()

def make_blb_sheet_aligned(file_name, sht_ix, bad_date_cols, target_indx, wrote_bk):
    """for each sheet in this file of bloomberg formatted data, all columns will
        be row aligned on each date. Must be xlsx workbook.
        file_name: name of book this operation is to be performed on
        sht_ix: sheet_ix
        bad_date_cols: list of columns of dates(as ints) which needs to be extended+next 2 cols
        target_indx: the dates that data_ix will now hold (as dt objects)
                     a strict superset of columns in bad_date_cols
                        (values only in bad_date_cols will be removed)
        wrote_bk: the openpyxl book to write to
        returns: the book that was opened, *MUST* be {{saved + closed}} by caller
        """
    sheet = wrote_bk.worksheets[sht_ix]
    for c in bad_date_cols:
        #Have to remove all values that aren't in target_indx
        valid_idx, d_col = list(zip(*[(ix, v.value)
                                for ix, v in enumerate(sheet[int_col(c)])
                                    if v.value in target_indx]
                                ))
        d_col = list(d_col)
        valid_idx = set(valid_idx)
        t_col = [v.value for ix, v in enumerate(sheet[int_col(c+1)])
                                    if ix in valid_idx]
        v_col = [v.value for ix, v in enumerate(sheet[int_col(c+2)])
                                    if ix in valid_idx]
        for ix, v in enumerate(target_indx):
            if ix >= len(d_col) or (d_col[ix] !=v and d_col[ix] != "#N/A N/A"):
                try:
                    print(f"inserted {v} @ {ix}, col {c}, was {d_col[ix]}")
                except:
                    print(f"inserted {v} @ end, col {c}")
                d_col.insert(ix, v)
                t_col.insert(ix, "#N/A N/A")
                v_col.insert(ix, "#N/A N/A")
            if len(d_col) > len(target_indx):
                pdb.set_trace()
        t_col += [ "#N/A N/A"]*(len(d_col) - len(t_col))#??, not sure why not same sz
        v_col += [ "#N/A N/A"]*(len(d_col) - len(v_col))

        #writeback uses excel locations, 1-indexed instead of 0
        for row in range(8, 8 + len(d_col)):
            sheet.cell(column=c+1, row=row).value = d_col[row-8]
            sheet.cell(column=c+2, row=row).value = t_col[row-8]
            sheet.cell(column=c+3, row=row).value = v_col[row-8]
    return wrote_bk

def sort_contracts(sec_names, latest_first = True):
    """Sorts Contracts by their date abreviations.
    Commodity must be space seperated from date, eg. 'C N1' """
    sec_mt , sec_yr = list(zip(*[re.findall("([a-zA-Z])(\d+)", i)[0]
                                 for i in sec_names]))

    assert all([len(i) == 2 or i in ('1', '0', '9') for i in sec_yr]), \
            f"Recheck date conversion, {sec_yr[:100]}"
    sec_yr = [21 if yr == '1' else
              20 if yr == '0' else
              19 if yr == '9' else
              int(yr)
                  for yr in sec_yr]
    month_abreviations = list(defs.con_month_abv.values())
    sec_mt = [month_abreviations.index(i)
                  for i in sec_mt]
    reformatted = [yr*100 + m
                       for yr,m in zip(sec_yr, sec_mt)]
    return [s for _,s in sorted(zip(reformatted, sec_names),
                                     key = lambda i: i[0],
                                     reverse= latest_first)]

def return_expired(sec_list, curve_prices, con_back = None):
    """Returns a pandas df with the expiry prices for each contract, with columns
    being the number of months since that expiry. Use for predicting future prices
    without concurrent trading, eg. prediciting August settle given June settle vs
    estimating current CL1 given current CL2.
    sec_list: list of pd.Series; used to select indexs for a given contract
    curve_prices: pd.Df, the front month's prices for a given contract are what
                    determines the return value
    con_back: the number of contracts ago to set columns as.
    if con_back is none returns NxN; else returns (N-con_back+1)xN
    """
    contract_abv = re.sub("\d+", "", sec_list[0].name)[:-1].strip()
    collection_date = max([max(i.index) for i in sec_list])
    expired_contract = [i for i in sec_list
                            if max(i.index)<collection_date ]
    sorted_contract = [i.name for i in sorted(expired_contract,
                                              key = lambda i: max(i.index),
                                              reverse = True)]

    prev_contract = {sorted_contract[i]: sorted_contract[i+1]
                                 for i in range(len(sorted_contract)-1)}
    prev_contract[sorted_contract[-1]] = 'last'
    con_2_expiry = {i.name: max(i.index) for i in expired_contract}
    con_2_expiry['last'] =  min(
                                next(
                                    filter(lambda i: i.name == sorted_contract[-1],
                                            sec_list)
                                      ).index
                                )
    #inclusive range of dates where front month
    con_range = {curr: (con_2_expiry[prev]+timedelta(1),
                        con_2_expiry[curr])
                      for curr, prev in prev_contract.items()}

    if con_2_expiry[sorted_contract[0]] < collection_date:
        #have front month prices that weren't used
        front_mo_con = sort_contracts([i.name
                                          for i in sec_list
                                            if max(i.index) == collection_date],
                                      latest_first = False
                                      )[0]
        con_range[front_mo_con] = (con_2_expiry[sorted_contract[0]] + timedelta(1),
                                   collection_date)
        sorted_contract.insert(0, front_mo_con)
    else:
        front_mo_con = ""

    def _expiry_prices(tick, out_sz=None):
        """"gets the expiry month prices for a given contract as np.array,
        tick is ticker of the expired contract.
        For contracts that have fewer trading days then out_sz,
        the prices in the *first* front month trading date will be used.
        This will be flipped for the very last, unexpired contract, which will
        map to first trading days of each contract(??this doesn't work??, grib)
        out_sz: size of df to be returned"""
        indxs = np.logical_and(curve_prices.index >= con_range[tick][0],
                                curve_prices.index <= con_range[tick][1])
        out = curve_prices.iloc[indxs, 0]#front month prices for this Commodity
        if out_sz:
            missing_sz = out_sz - len(out)
            if missing_sz > 0:
                #contract has fewer trading days then some contract after it
                return np.concatenate((out.values,
                                      np.repeat(out.iloc[-1], missing_sz)))
            else:
                if tick != front_mo_con:
                    return out[:out_sz].values
                else:
                    return out[-out_sz:].values
        else:
            return out.values

    def _make_filler(t,ix):
        """NAs to add to 'back' of df for contracts that weren't traded over the entire period.
        eg. to the back of CLZ18 for the dates beyond (Dec 18 - start of data)
        when this wasn't traded for the entire length of backtest
        Needs to be transposed"""
        if con_back:#no filler
            return [[np.nan]*len( _expiry_prices(t)) for _ in range(0)]
        else:
            return [[np.nan]*len( _expiry_prices(t)) for _ in range(ix)]

    def _make_index():
        "Makes the index for ticker t of datetime objects"
        #each list(filter) will be same length as
        if con_back and con_back > 1:
            dt_index =  [ix for tick in sorted_contract[:-con_back+1]
                             for ix in filter(lambda i: i >= con_range[tick][0]
                                                    and i <= con_range[tick][1],
                                              curve_prices.index)]
        else:
            dt_index =  [ix for tick in sorted_contract
                             for ix in filter(lambda i: i >= con_range[tick][0]
                                                    and i <= con_range[tick][1],
                                              curve_prices.index)]
        named_index = [f"{t} {i} before expiry"
                            for t in sorted_contract
                                for i in range(len(_expiry_prices(t)))]
        tuples = list(zip(dt_index, named_index))
        multi_indx = pd.MultiIndex.from_tuples(tuples,
                                               names=['Dates', 'Description'])

        axis_unique = len(np.unique(dt_index)) == len(dt_index)
        assert axis_unique, "datetime axis is not unqiue, repeated dates"
        return dt_index

    if con_back and con_back > 1:#filters partial NA rows
        zero_ago_tickers = enumerate(sorted_contract[:-con_back+1])
    else:
        zero_ago_tickers = enumerate(sorted_contract)
    num_cols = con_back or len(sorted_contract)

    expired_curve = pd.DataFrame(
                        np.concatenate(
                            [np.stack(
                                [_expiry_prices(j,
                                                out_sz = len( _expiry_prices(t)))
                                     for j in sorted_contract[ix:ix+num_cols]]
                                + _make_filler(t,ix)
                                 ).T
                                    for ix, t in zero_ago_tickers #row chunks
                            ]),
                        index = _make_index(),
                        columns = [f"{contract_abv} {i}Ago" #not all contracts 1Mo
                                   for i in range(num_cols)]
                        )
    return expired_curve

def process_macroTrendsnet(file = None):
    """Return df of all excel sheets from macrotrend's in folder, are spot prices
    if file != none then just that file's series.
    NOTE: MacroTrends changes contracts in the middle of the month(15th?),
    at a different time than bloomberg/WTI. There's a difference on ~15-20 every month.
    I think macrotrends still uses the old WTI expiry dates?(Then why holds for 2019)?"""
    os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio\\Data\MacroTrends")
    def _1_macroTrends_sht(file):
        name = file.split("-")[0]
        df = pd.read_csv(file,
                         header = 15,
                         index_col = 0,
                         parse_dates = True,
                         dtype={'value':np.float64}).iloc[:,0]
        # df.columns = [name]#no columns since series
        df.name = name
        if name in ('wheat', 'corn'):
            df *= 100
        return df[~df.isna()]

    if file is not None:
        return _1_macroTrends_sht(file
                                  ).sort_index(ascending = False)
    else:
        out = {}
        for file in os.listdir():
            name = file.split("-")[0]
            out[name] = _1_macroTrends_sht(file)
        return pd.concat(out,
                         axis=1,
                         join='outer',
                         sort=True
                         ).sort_index(ascending = False)

#%Functions to call directly
def get_blb_excel(prices_file, already_formatted = False):
    """Reads in data from an excel workbook with many sheets of Bloomberg formated
    historical future's prices, with Top Row Dates aligned.
    Returns Pandas DF's
    """
    if not already_formatted:
        print(f"Will be long, is reformatting, make sure {prices_file} is closed")
        make_blb_book_row_aligned(prices_file)#modifies book

    xl_bk = xlrd.open_workbook(prices_file)
    commodities = xl_bk.sheet_names()
    security_l = []
    curve_prices_l = []
    expired_curves_d = {}

    for sht_ix, name in enumerate(commodities):
        b = xl_bk.sheet_by_index(sht_ix)
        com_maturities = [i for i in b.row_values(0)[1::4] if i]
        sz = len([i for i in b.row_values(6)])
        date_cols = [[datetime(*xlrd.xldate_as_tuple(j,0))
                            for j in b.col_values(i)[7:]
                        if j]
                            for i in range(0,sz,4)]
        dates = sorted(set([i
                            for c in date_cols
                                for i in c]),
                        reverse = True)
        curve_prices = pd.DataFrame(
                                np.array(
                                    [pd.to_numeric(b.col_values(i)[7:],
                                                   errors = 'coerce')
                                           for i in range(2,sz,4)]
                                        ).T,
                                     columns = com_maturities,
                                     index = dates,
                                     )
        curve_prices = curve_prices.dropna(how='all', axis=0
                                           ).dropna(how='all', axis=1)
        curve_prices_l += [curve_prices]#.dropna()

        ##individual_securities
        sec_df = pd.DataFrame(list(zip(*[b.col_values(i)[7:]
                                                for i in range(1,sz,4)]
                                                )) )
        securities = np.unique(sec_df.values)
        securities = securities[~np.isin(securities, ('', '#N/A N/A'))]

        def get_security(s):
            "returns prices for 1 single future"
            rows, _ = np.where(sec_df == s)
            #get all of row, column not just those indexes.
            df = pd.Series(data = curve_prices.values[sec_df == s],
                      index = curve_prices.index[rows],
                      name = s
                      ).dropna()
            return df[~pd.isna(df.index)]
        sec_list = list(filter(lambda i: len(i) > 0,
                               [get_security(s) for s in securities]))
        security_l += sec_list

        ##expired securities
        contract_abv = re.sub("\d+", "", sec_list[0].name)[:-1].strip()
        expired_curve = return_expired(sec_list,
                                        curve_prices)
        expired_curves_d[contract_abv] = expired_curve

    #make dataframes
    curve_prices_df = pd.concat(curve_prices_l,
                                axis=1,
                                join='outer',
                                sort=True
                                ).iloc[::-1]#largest axis up top
    # curve_prices_df.dropna(how='all', axis=0, inplace=True)
    curve_prices_df.columns = [i.replace("COMB", "").replace("Comdty", "").replace(" ", "")
                for i in curve_prices_df.columns]#eg CL 1
    curve_prices_df.dropna(how='all', axis=1, inplace=True)
    securities_df = pd.concat(security_l,
                              axis=1,
                              join='outer',
                              sort=True).iloc[::-1]
    #sizes are off since Some columns have date's with prices, but no ticker??
    securities_df.dropna(how='all', inplace=True)
    expired_curves_df = pd.concat(expired_curves_d.values(),
                                  axis=1,
                                  join='outer').iloc[::-1]
    return curve_prices_df, securities_df, expired_curves_df

def save_struct(struct, name):
    "handler to pickle data"
    d = os.getcwd()
    os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio\\Data\\pickled_data")
    if isinstance(struct, pd.DataFrame) or isinstance(struct, pd.Series):
        struct.to_pickle(f'{name}.p')
    else:
        with open(f'{name}.p', 'wb') as file:
            pickle.dump(struct,  file)
    os.chdir(d)

def load_struct(name):
    d = os.getcwd()
    os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio\\Data\\pickled_data")
    try:
        with open(f'{name}.p', 'rb') as file:
            return pickle.load(file)
    except:
        return pd.read_pickle(f'{name}.p')
    os.chdir(d)

def reprocess_struct(name):
    "Deletes pickled data so data_handler will reload it"
    d = os.getcwd()
    os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio\\Data\\pickled_data")
    try:
        os.remove(f"{name}.p")
    except Exception as e:
        print(f"{name} alread removed", e)
    try:
        exec(f"del {name}")
    except:
        pass
    data_handler(data_structs = [name])
    os.chdir(d)


def data_handler(save_data = False,
                 data_structs = ('curve_prices_df',
                                'securities_df',
                                'expired_curves_df',
                                'cme_df',
                                'eia_bio_df',
                                'historic_front_month',
                                'retail_diesel2')):
    """"MODIFIES GLOBALS; by assigning values to data_structs
    All purpose data handler, uses subfolder 'picked_data'.
    Gets all seperated data sources
    """
    if save_data:
        for struct_name in data_structs:
            try:
                struct = eval(struct_name)
                save_struct(struct, struct_name)
            except Exception as e:
                print(e)
        return None
    else:#load data
        for struct_name in data_structs:
           try:
               exec(f"global {struct_name}", globals())
               exec(f"globals()[struct_name] = load_struct(struct_name)")
           except Exception as e:
               print(f"""Haven't pickeled {struct_name}, Processing data Now.\n""", e, "\n\n")

               if struct_name in ('curve_prices_df', 'securities_df', 'expired_curves_df'):
                   os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio\\Data")
                   prices_file = "16.16 Historical Commodity Price Data.xlsx"
                   (curve_prices_df,
                    securities_df,
                    expired_curves_df) = get_blb_excel(prices_file,
                                                      already_formatted  = True)

                   save_struct(curve_prices_df, 'curve_prices_df')
                   save_struct(securities_df, 'securities_df')
                   save_struct(expired_curves_df, 'expired_curves_df')

               elif struct_name == 'cme_df':
                   cme_df = cme_scrapper.make_cme_df(cme_scrapper.get_all_cme_prx())
                   save_struct(cme_df,'cme_df')

               elif struct_name == 'eia_bio_df':
                   #https://www.ers.usda.gov/data-products/us-bioenergy-statistics/
                   os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio\\Data")
                   eia_bio_df = cme_scrapper.eia_renewable_table(table=17)
                   save_struct(eia_bio_df, 'eia_bio_df')

               elif struct_name == 'historic_front_month':
                   historic_front_month = process_macroTrendsnet()
                   save_struct(historic_front_month, 'historic_front_month')

               ##unused
               elif struct_name == 'cotton':
                   cotton = process_macroTrendsnet(
                               file = 'cotton-prices-historical-chart-data.csv')
                   save_struct(cotton, 'cotton')

               elif struct_name == 'retail_diesel2':
                   #https://www.eia.gov/dnav/pet/hist/LeafHandler.ashx?n=pet&s=emd_epd2d_pte_nus_dpg&f=w
                   os.chdir("C:\\Users\\student.DESKTOP-UT02KBN\\Desktop\\Stone_Presidio\\Data")
                   retail_diesel2 = pd.read_excel('weekly_US_No2_Retail_diesel.xls',
                                              sheet_name = 'Data 1',
                                              header=2,
                                              index_col=0)
                   retail_diesel2 = retail_diesel2.iloc[::-1]
                   retail_diesel2.columns = ["No 2 Diesel Retail"]
                   save_struct(retail_diesel2, 'retail_diesel2')

               exec(f"global {struct_name}", globals())
               exec(f"globals()[struct_name] = load_struct(struct_name)")
        #not nessisary
        return [eval(i) for i in data_structs]

#Makes pyflakes behave
(curve_prices_df,
securities_df,
expired_curves_df,
cme_df,
eia_bio_df,
historic_front_month,
retail_diesel2) = data_handler(save_data = False)

all_data =load_struct('all_data')
#%% Predicting comodity pries
from sklearn.experimental import enable_iterative_imputer
from sklearn.impute import IterativeImputer, KNNImputer
#iterativeIMputer doesn't have convergence guarantees
from sklearn.impute import SimpleImputer
from sklearn import linear_model

seed = 0

def bridge_imputer(df, trend_model="date_reg", seed = seed, std_range = 50, print_r2=False):#x1, x2, sd, ln
    """ takes pandas df or Series with index *Reverse* sorted index
    Returns pd.df with na values imputed from a brownian bridge.
    use_trend: Use linear model in addition to residuals
    std_range: Size of index window to take std over; excluding NAs, before period start
    """
    np.random.seed(seed)
    if isinstance(df, pd.Series):
        df = df.to_frame()
    else:
        df = df.copy()

    pre_std = df.std()
    for c in df.columns:
        x = df[c]
        f_indx = x.first_valid_index()
        f_ix = len(x[:f_indx])
        l_indx = x.last_valid_index()
        l_ix = len(x[:l_indx])
        #time varying Var + std?
        if trend_model == 'start_stop':
            raise Exception("Makes very Bad predictions")
            m = (x[f_indx] - x[l_indx])/(f_ix - l_ix)
            trend = [np.nan]*f_ix \
                    + [m*i if v==v else np.nan
                       for i,v in enumerate(x.iloc[f_ix:l_ix+1])] \
                    + [np.nan]*(len(x) - l_ix - 1)
            detrend = x - trend#np.std(x - trend)
            if print_r2:
                z = x.dropna()
                m = sum(z)/len(z)
                t = [q + m for q in trend if q==q]
                rss = sum([(zz-tt)**2 for zz,tt in zip(z,t)])
                ss = sum([(zz-m)**2 for zz in z])
                print(c, "Start Stop R^2 = ", 1 - rss/ss)
            if  np.nanvar(x -trend)/np.nanvar(x) > 1.05:
                print(f"Trend for {c} increased var:",
                      f"{np.nanvar(x -trend):.3f} vs. {np.nanvar(x):.3f}.")
                      # f"corr coef: {np.cov(x*trend)/np.sqrt(np.nanvar(x)*np.nanvar(trend))}")
            if detrend.iloc[0] != detrend.iloc[0]:
                detrend.iloc[0] = x[f_indx] - m*f_ix
            if detrend.iloc[-1] != detrend.iloc[-1]:
                detrend.iloc[-1] = x[l_indx] + m*(len(x) - l_ix - 1)
        elif trend_model == 'date_reg':
            model = None
            cor = 0
            season = None
            for j in [3,5,8]:#turn around season varies by commodity; Spring vs. Fall or Summer(start or end) vs. Winter
                dates = np.array([(i.year, abs(i.month + i.day/i.month -j))
                                    for i in  x.index
                                    if x[i] == x[i]])
                temp = x.dropna()

                temp_mod= linear_model.LinearRegression().fit(dates,)
                if temp_mod.score(dates, x.dropna()) > cor:
                    model = temp_mod
                    season = j
                    cor = temp_mod.score(dates, x.dropna())
            if print_r2:
                print(c, f"R^2 = {cor:.2f}", season)
            detrend = pd.Series(x.index,
                              index = x.index
                              ).apply(lambda i:
                                         x[i] if x[i]==x[i]
                                             else model.predict(
                                                 np.array([[i.year,
                                                            abs(i.month
                                                                + i.day/i.month
                                                                -season)]]))[0])
            if detrend.iloc[0] != detrend.iloc[0]:
                i = x.index[0]
                detrend.iloc[0] =  model.predict(np.array([[i.year,
                                                      abs(i.month-season)]]))
            if detrend.iloc[-1] != detrend.iloc[-1]:
                i = x.index[-1]
                detrend.iloc[-1] =  model.predict(np.array([[i.year,
                                                       abs(i.month-season)]]))
        else:
            if detrend.iloc[0] != detrend.iloc[0]:
                detrend.iloc[0] = x[f_indx]
            if detrend.iloc[-1] != detrend.iloc[-1]:
                detrend.iloc[-1] = x[l_indx]
            detrend = x - sum(x)/len(x)

        valid_ixs = np.where(~x.isna())[0]
        bridge_ixs = [(i,j) for i,j in zip(valid_ixs[:-1],
                                           valid_ixs[1:])
                        if i +1 < j]

        temp = detrend.dropna()
        m1 = sum(temp[:std_range])
        m2 = sum(temp[:std_range]**2)
        ts_ix = 0#inclusive
        te_ix = std_range-1#inclusive
        for s, e in bridge_ixs:
            s_indx = x.index[s]#not in temp axis
            try:
                s_ix = temp.index.get_loc(s_indx) + 1 #exclusive
            except:
                pdb.set_trace()

            n_added = max(len(temp.iloc[te_ix:s_ix])-1, 0)
            m1 += sum(temp.iloc[te_ix+1:s_ix])  \
                - sum(temp.iloc[ts_ix:ts_ix+n_added])
            m2 += sum(temp.iloc[te_ix+1:s_ix]**2)  \
                - sum(temp.iloc[ts_ix:ts_ix+n_added]**2)
            ts_ix += n_added
            te_ix += n_added
            # te_indx = temp.index[te_ix]
            date_range = (temp.index[ts_ix] - temp.index[te_ix]).days#unbiased est
            #issue of calc std over indexs, but then use for sd over dates, if gaps vary in size
            sd = np.sqrt((m2/std_range - (m1/std_range)**2))
            # if  c == 'Retail Biodiesel' and s < date_range:#and s%3 == 0
            #     print(c, [f"{i:.2f}" for i in (sd, temp.iloc[ts_ix:te_ix+1].std(),
            #           m1, sum(temp.iloc[ts_ix:te_ix+1]),
            #           m2, sum(temp.iloc[ts_ix:te_ix+1]**2))])
            # if abs(m1 - sum(temp.iloc[ts_ix:te_ix+1])) > 0.01:
            #     ts_ix -= n_added
            #     te_ix -= n_added
            #     te_indx = temp.index[te_ix]
            #     m1 -= sum(temp.iloc[te_ix+1:s_ix])  \
            #         - sum(temp.iloc[ts_ix:ts_ix+n_added])
            #     m2 -= sum(temp.iloc[te_ix+1:s_ix]**2)  \
            #         - sum(temp.iloc[ts_ix:ts_ix+n_added]**2)
            #     pdb.set_trace()

            a,b = detrend.iloc[s], detrend.iloc[e]
            ln = len(detrend.iloc[s+1:e])
            #assumes equal var over dates w/o indexs
            days_passed = [(i-j).days for i,j in zip(x.index[s+1:e],
                                                     x.index[s+2:e+1])]
            mx, mn = max(x) + 2*sd, min(x) - 2*sd
            wiener_proc = np.cumsum([np.random.normal(0, sd*np.sqrt(d))
                                     for d in days_passed])
            detrend.iloc[s+1:e] = [a + (b-a)*i/ln + w - wiener_proc[-1]*i/ln
                             for i, w in enumerate(wiener_proc)]
        df.loc[:,c] = x
    std_ratio = df.std()/pre_std
    if any([i>1.15 or i < 0.85 for i in std_ratio]):
        print("The following data series changed SD by more than 15%:",
              [(c,f"{i*100 - 100}% change")
                for i,c in zip(std_ratio, df.columns)
                    if i>1.15 or i < 0.85])
    return df

# all_data =load_struct('all_data')
# isint = lambda i: i in '0123456789'
# all_data = all_data.loc[:, [n for n in all_data.columns if ((isint(n[-1]) and n[-1] == '1') or (not isint(n[-1])))]]
z = bridge_imputer(all_data['Retail Biodiesel'], trend_model = 'date_reg', print_r2=True)
# print("\n\n\n", z.isna().sum())
# all_data =load_struct('all_data')
# bridge_imputer(all_data, trend = 'date_reg',print_r2=True)


# z = all_data.loc[:,'Retail Biodiesel']
# pre_std = all_data.iloc[:3000,:].std()

# all_data= bridge_imputer(all_data, trend = False)
#%%
# z = all_data
all_data =load_struct('all_data')

std_ratio = z.iloc[:4000,:].std()/all_data.iloc[:4000,:].std()
if any([i>1.15 or i < 0.85 for i in std_ratio]):
    print("The following data series changed STD by more than 15%:",
          [(c,i)
            for i,c in zip(std_ratio, all_data.columns)
                if i>1.15 or i < 0.85])
#%%
import matplotlib.pyplot as plt

# plt.scatter(all_data.loc[:,'Retail Biodiesel'].index,
#               all_data.loc[:,'Retail Biodiesel'],
#               s = 3,
#               marker = '_')
plt.acorr(z.dropna(), maxlags=400, normed=True)
plt.show()
#%%
fig, ax = plt.subplots()
# z = all_data.loc[:,'Retail Biodiesel']
indx = z.index[4600]
ax.scatter(z[:indx].index,#'Retail Biodiesel'].index,
              z[:indx],#'Retail Biodiesel'],
              s = 0.1)
# all_data = load_struct('all_data')
ax.scatter(all_data.loc[:indx,'Retail Biodiesel'].index,
              all_data.loc[:indx,'Retail Biodiesel'],
              s = 3,
              marker = '_')
fig.show()
#%%
plt.plot(imputer.fit_transform(all_data.loc[:,'Retail Biodiesel'].values.reshape(-1,1)))
plt.show()



    #%%
def preprocessing(months_back = 15,
                  impute = KNNImputer(n_neighbors=10, weights='distance').fit_transform):
    """Do all preprocessing to return X_train, Y_train, X_test, y_test.
    combine data, Remove all NAs from data.
    months_back: how far back to keep expires prices for, approx.
    start_date: The oldest date for which prices will be kept"""
    #No NA Valid NA, or Valid, NA, Valid in curve_prices:  ((expired_curves_df.iloc[:-2,:] == expired_curves_df.iloc[:-2,:]) & (expired_curves_df.iloc[1:-1,:] != expired_curves_df.iloc[1:-1,:]) & (expired_curves_df.iloc[2:,:] == expired_curves_df.iloc[2:,:])).sum().sum()
    #But NAs of prices before started trading/didn't trade that far back? Invert regression
    end_date = securities_df.index[0]
    min_gap = timedelta(30*months_back)
    start_date = min_gap + securities_df.index[-1]

    #BLB sheet data
    con_abv = set([i[:2] for i in securities_df.columns])
   #expired contracts only
    con_d = {c: sort_contracts(
                    [i[2:]
                     for i in securities_df.columns
                        if i[:2] == c
                        and np.isnan(securities_df.loc[end_date, i])
                        ])
                for c in con_abv}
    #last trading date for each contract
    con_d_last = {k: [securities_df.index[
                        next(i
                           for i,x in enumerate(securities_df[k+con])
                               if x == x)#filters NAs
                               ]
                        for con in l]
                  for k,l in con_d.items()}
    #some contracts have changed in length, why n^2
    con_numCon = {k: next((i
                          for i in range(1, months_back)
                           if min([e-s
                                   for s,e in zip(l[i:], l[:-i])])
                              > min_gap),
                          months_back)
                      for k,l in con_d_last.items()}
    columns = [f"{contract_abv.strip()} {i}Ago"
                   for contract_abv, m_back in con_numCon.items()
                       for i in range(1,m_back)]
    #Which ever length of columns is smallest, fewer Nans
    row_ix = max([next(ix
                     for ix, v in enumerate(reversed(expired_curves_df[c]))
                         if v == v)
                for c in columns])
    last_row = expired_curves_df.index[-row_ix]

    curve_columns_d = {c:[i
                            for i in curve_prices_df.columns
                             if re.sub("\d+", "", i) == c.strip()]
                           for c in con_abv}
    curve_prices_d = {c: pd.DataFrame(
                            impute(
                                curve_prices_df.loc[:,curve_columns_d[c]]
                                ),
                            columns = curve_columns_d[c],
                            index = curve_prices_df.index
                            )
                      for c in con_abv}
    #filter dates after make dict so can be included in expired_curves
    curve_df = pd.concat(curve_prices_d.values(),
                                axis=1,
                                join='outer',
                                sort=True).iloc[::-1]
    curve_df = curve_df[curve_df.index >= last_row]
    #rebuild expiry_prices to use imputed values from curve_prices
    sec_list = [securities_df[i].dropna() for i in securities_df.columns]
    sec_list_d = {c: list(filter(lambda i: i.name[:2] == c, sec_list))
                      for c in con_abv}
    #sec_list values not used so don't need to update
    expired_df = pd.concat([return_expired(sec_list_d[c],
                                                  curve_prices_d[c],
                                                  con_back = con_numCon[c]
                                                  ).iloc[:, 1:]#drop 0Ago values
                                   for c in con_abv],
                                  axis=1,
                                  join='outer').iloc[::-1]
    #outerjoin leaves Nas since different indexs; since different contracts expire at different times.
    expired_df.iloc[:,:] = imputer.fit_transform(expired_df)#grib?

    #macroTrends: only take unique commoditie's columns, after start date
    front_mth = historic_front_month[['cotton']].dropna()
    front_mth = front_mth[np.logical_and(end_date >= front_mth.index,
                                         front_mth.index >= start_date)]

    #EIA (bio-)diesel
    r_diesel2 = retail_diesel2[np.logical_and(end_date >=retail_diesel2.index,
                                         retail_diesel2.index >= start_date)]
    biodiesel = eia_bio_df.loc[np.logical_and(end_date >= eia_bio_df.index,
                                              eia_bio_df.index >= start_date),
                               'Retail Biodiesel']
    biodiesel.index = [next(j
                            for j in reversed(r_diesel2.index)
                            if j >= i)
                        for i in biodiesel.index]
    retail_diesels = r_diesel2.join(biodiesel,
                                    how='left',
                                    sort = True
                                    ).fillna(method='ffill')
    all_data = pd.concat([curve_df,
                          expired_df,
                          front_mth,
                          retail_diesels],
                         join='outer',
                         axis = 1,
                         sort=True).iloc[::-1]

    return all_data

all_data = preprocessing(impute = bridge_imputer)
#Difference of what dates should have data
# [i for i in all_data['SM 1Ago'][all_data['SM 1Ago'].isna()].index if i in curve_df.index]
#%%
from sklearn.model_selection import train_test_split
def make_prediction_df():
    front_cols = [i for i in curve_prices_df.columns
                 if '1 ' in i and '11' not in i and '21' not in i]
    rest_cols = [i for i in curve_prices_df.columns if i not in front_cols]

    X = curve_prices_df.loc[:,rest_cols]
    Y = curve_prices_df.loc[:,front_cols[0]]
    X_train, X_test, y_train, y_test = train_test_split(X, Y, random_state=0)


#%%
from sklearn.pipeline import Pipeline
from sklearn.linear_model import LinearRegression
from sklearn.preprocessing import StandardScaler

pipe = Pipeline([('impute', SimpleImputer()),
                     ('scaler', StandardScaler()),
                     ('linear', LinearRegression())])

pipe.fit(X_train, y_train)



#%%


#%% Make Price Distribution Graphs from future's
from scipy import integrate

futures = [i.replace("COMB", "").replace("Comdty", "").replace(" ", "")
            for i in curve_prices_df]#eg CL 1
futures_ab = set([re.sub("\d+", "",i)
                    for i in futures])#eg CL

# month_rng = sorted(set([datetime(d.year, d.month, 1)
#                         for d in securities_df.index]))
# month_abvs = sort_contracts([f"{defs.int_month_abv[i.month]}{str(i.year)[-2:]}"
#                               for i in month_rng])

hist_pct = hist_val.pct_change(periods = 252)#6mo out
hist_vol = hist_pct.std(axis=0)

for ab in defs.abv_cme_url:
    if ab:
        tick = ab + "1"
        ix = futures.index(tick)
        base_ix = hist_val.iloc[:,ix].first_valid_index()
        base_prx = hist_val.loc[base_ix, tick]
        dist = ss.lognorm(loc = 0, scale = 1, s = hist_vol.values[ix])

        x = np.linspace(hist_pct.iloc[:,ix].min(),
                        hist_pct.iloc[:,ix].max(),
                        100)*2

        fig, ax =plt.subplots()
        x_axis = [i for i in  base_prx * (1 + x) if i > 0]

        probs = [integrate.quad(lambda x: dist.pdf(x/base_prx)/base_prx, i,j)[0]
                         for i,j in zip(x_axis[:-1], x_axis[1:])]
        probs = [100*i/sum(probs) for i in probs]

        plt.bar(x_axis[:-1], probs, label="Probabilities")
        plt.legend()
        plt.title(f"Probability of ending prices of {hist_vol.index[ix]} in 12mo")
        xformatter = mticker.FormatStrFormatter('$%.0f')
        ax.xaxis.set_major_formatter(xformatter)
        yformatter = mticker.FormatStrFormatter('%.1f%%')
        ax.yaxis.set_major_formatter(yformatter)
        plt.show()


#plt.plot(x_axis, dist.pdf(x), label='dist')
#plt.yscale('log')